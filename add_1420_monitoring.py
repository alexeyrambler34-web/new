from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Users\HONOR\Documents\Codex\2026-06-18\files-mentioned-by-the-user-17062026")
OUTPUTS = BASE / "outputs"
INPUT = OUTPUTS / "Обзор цен 19062026 только цены мониторинг.xlsx"
OUTPUT = OUTPUTS / "Обзор цен 19062026 только цены мониторинг авто 1420.xlsx"

STAMP = "19.06.2026 14:20"
RED_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
NO_FILL = PatternFill(fill_type=None)


def note(marketplace, title):
    return f"{marketplace}, {STAMP} — {title}"


data = {
    4: {
        "ГК": (11450, note("Ozon", "GERVENT Система вытяжной вентиляции для металлочерепицы, коричневый")),
        "РВТ": (5805, note("Ozon", "GERVENT Нанодефлектор РВТ-160, коричневый")),
        "Static/Колпак": (7417, note("Ozon", "GERVENT Static Комплект кровельной вентиляции, темно-коричневый")),
    },
    6: {
        "ГК": (11403, note("Wildberries", "GERVENT Готовый комплект, коричневый")),
        "РВТ": (6707, note("Wildberries", "GERVENT Нанодефлектор / комплект кровельной вентиляции, коричневый")),
    },
    7: {
        "ГК": (9231, note("Wildberries", "GERVENT Комплект вентиляции с ротационной турбиной, серый графит")),
        "РВТ": (9481, note("Wildberries", "GERVENT Ротационная вентиляционная турбина, зеленый")),
        "Static/Колпак": (8097, note("Wildberries", "GERVENT Комплект вентиляции Static, серый")),
    },
    8: {
        "ГК": (12694, note("Яндекс Маркет", "Нанодефлектор с вентвыходом на плоскую кровлю, серый графит")),
        "РВТ": (5760, note("Wildberries", "GERVENT Дефлектор ротационный, черный")),
        "Static/Колпак": (9762, note("Wildberries", "GERVENT Комплект вентиляции Static, коричневый")),
    },
    13: {
        "ГК": (11720, note("Ozon", "GERVENT Комплект кровельной вентиляции для металлочерепицы Монтеррей, зеленый")),
        "ВВ": (3432, note("Ozon", "GERVENT Вентиляционный выход из ABS-пластика ND, серый")),
        "Static/Колпак": (5947, note("Wildberries", "GERVENT Комплект вентиляции Static для фальцевой кровли, черный")),
    },
    18: {
        "ГК": (9555, note("Ozon", "Готовый комплект изолированный для кровли с профилем Монтеррей, серый графит")),
        "РВТ": (3335, note("Яндекс Маркет", "Ротационная вентиляционная турбина РВТ 160 мм, серый графит")),
        "ВВ": (2373, note("Яндекс Маркет", "Вентиляционный выход изолированный 125/160 мм, серый графит")),
        "Static/Колпак": (7056, note("Ozon", "Готовый комплект Static для кровли с профилем Монтеррей, серый графит")),
    },
    21: {
        "ГК": (10690, note("Ozon", "Комплект вентиляции для плоской кровли, вытяжка для бани, черный")),
        "РВТ": (5343, note("Яндекс Маркет", "Ротационная вентиляционная турбина Нанодефлектор D160, коричневый")),
    },
    23: {
        "Static/Колпак": (7450, note("Wildberries", "GERVENT Комплект кровельной вентиляции D160, коричневый")),
    },
    45: {
        "РВТ": (7450, note("Wildberries", "GERVENT Комплект для металлочерепицы Монтеррей, коричневый")),
    },
}


def find_link_start(ws):
    for col in range(1, ws.max_column + 1):
        value = ws.cell(3, col).value
        if value and "ссылка на магазин" in str(value).lower():
            return col
    raise RuntimeError("Не найдены колонки ссылок")


def copy_cell_format(src, dst):
    dst._style = copy(src._style)
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def main():
    wb = load_workbook(INPUT)
    ws = wb.active

    link_start = find_link_start(ws)
    link_count = ws.max_column - link_start + 1

    saved_links = []
    for row in range(1, ws.max_row + 1):
        row_links = []
        for col in range(link_start, ws.max_column + 1):
            cell = ws.cell(row, col)
            row_links.append(
                {
                    "value": cell.value,
                    "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
                    "style": copy(cell._style),
                    "font": copy(cell.font),
                    "fill": copy(cell.fill),
                    "border": copy(cell.border),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                    "protection": copy(cell.protection),
                    "comment": copy(cell.comment) if cell.comment else None,
                }
            )
        saved_links.append(row_links)

    ws.insert_cols(link_start, 4)
    new_start = link_start
    old_block_start = link_start - 4
    categories = ["ГК", "РВТ", "ВВ", "Static/Колпак"]

    for offset in range(4):
        src_col = old_block_start + offset
        dst_col = new_start + offset
        ws.column_dimensions[get_column_letter(dst_col)].width = ws.column_dimensions[get_column_letter(src_col)].width
        for row in range(1, ws.max_row + 1):
            copy_cell_format(ws.cell(row, src_col), ws.cell(row, dst_col))
            ws.cell(row, dst_col).value = None
            ws.cell(row, dst_col).comment = None
            ws.cell(row, dst_col).fill = copy(NO_FILL)

    ws.cell(1, new_start).value = datetime(2026, 6, 19, 14, 20)
    ws.cell(1, new_start).number_format = "dd.mm.yyyy hh:mm"
    ws.cell(2, new_start).value = "Цена, руб"
    for offset, category in enumerate(categories):
        ws.cell(3, new_start + offset).value = category

    own_gk = data[4]["ГК"][0]
    col_by_category = {category: new_start + i for i, category in enumerate(categories)}
    for row, values in data.items():
        for category, payload in values.items():
            value, comment_text = payload
            cell = ws.cell(row, col_by_category[category])
            cell.value = value
            cell.comment = Comment(comment_text, "Codex")
            cell.fill = copy(NO_FILL)

            if category == "ГК" and row != 4 and value < own_gk:
                cell.fill = copy(RED_FILL)
            elif category == "РВТ" and value < 5750:
                cell.fill = copy(RED_FILL)
            elif category == "Static/Колпак" and value < 7900:
                cell.fill = copy(RED_FILL)

    restored_link_start = new_start + 4
    for row_index, row_links in enumerate(saved_links, start=1):
        for offset, saved in enumerate(row_links):
            cell = ws.cell(row_index, restored_link_start + offset)
            cell.value = saved["value"]
            cell.hyperlink = saved["hyperlink"]
            cell._style = copy(saved["style"])
            cell.font = copy(saved["font"])
            cell.fill = copy(saved["fill"])
            cell.border = copy(saved["border"])
            cell.alignment = copy(saved["alignment"])
            cell.number_format = saved["number_format"]
            cell.protection = copy(saved["protection"])
            cell.comment = saved["comment"]

    for offset in range(link_count):
        old_col_letter = get_column_letter(link_start + offset)
        new_col_letter = get_column_letter(restored_link_start + offset)
        ws.column_dimensions[new_col_letter].width = ws.column_dimensions[old_col_letter].width

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
