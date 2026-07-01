from copy import copy
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(r"C:\Users\HONOR\Documents\Codex\2026-06-18\files-mentioned-by-the-user-17062026")
INPUT = BASE / "outputs" / "Обзор цен 19062026 только цены мониторинг авто 1420.xlsx"
OUTPUT = BASE / "outputs" / "Обзор цен 21062026 только цены мониторинг.xlsx"

src_wb = load_workbook(INPUT)
src_ws = src_wb.active
dst_wb = load_workbook(OUTPUT)
dst_ws = dst_wb.active

old_link_col = None
for col in range(1, src_ws.max_column + 1):
    value = src_ws.cell(3, col).value
    if isinstance(value, str) and "ссылка" in value.lower():
        old_link_col = col
        break

if old_link_col is None:
    raise RuntimeError("Не удалось найти ссылки в исходном файле.")

new_block = range(old_link_col, old_link_col + 4)
new_link_col = old_link_col + 4

for row in range(1, dst_ws.max_row + 1):
    for col in new_block:
        cell = dst_ws.cell(row, col)
        cell.hyperlink = None
        cell._hyperlink = None
        if row >= 4 and isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
            cell.value = None
            cell.comment = None

for offset in range(3):
    src_col = old_link_col + offset
    dst_col = new_link_col + offset
    dst_ws.column_dimensions[dst_ws.cell(1, dst_col).column_letter].width = src_ws.column_dimensions[src_ws.cell(1, src_col).column_letter].width
    for row in range(1, src_ws.max_row + 1):
        src = src_ws.cell(row, src_col)
        dst = dst_ws.cell(row, dst_col)
        dst.value = src.value
        dst.hyperlink = None
        dst._hyperlink = None
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        target = src.hyperlink.target if src.hyperlink else None
        if target:
            dst.hyperlink = target

dst_wb.save(OUTPUT)
print(OUTPUT)
