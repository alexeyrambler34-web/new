from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


BASE = Path(r"C:\Users\HONOR\Documents\Codex\2026-06-18\files-mentioned-by-the-user-17062026\outputs\Обзор цен 21062026 только цены мониторинг.xlsx")
OUT = Path(r"C:\Users\HONOR\Documents\Codex\2026-06-18\files-mentioned-by-the-user-17062026\outputs\Обзор цен 25062026 только цены мониторинг.xlsx")

DATE_TEXT = "25.06.2026"
FIX_TEXT = "25.06.2026 11:13-11:21"
LIGHT_RED = PatternFill(fill_type="solid", fgColor="F4CCCC")
NO_FILL = PatternFill(fill_type=None)


def note(marketplace: str, title: str) -> str:
    return f"{marketplace}, {FIX_TEXT} — {title}"


DATA = {
    4: {
        "ГК": (11049, note("Wildberries", "GERVENT комплект вентиляции / готовый комплект, серый графит")),
        "РВТ": (5821, note("Wildberries", "GERVENT Нанодефлектор РВТ-160, коричневый/серый графит")),
        "Static/Колпак": (7373, note("Wildberries", "GERVENT Static комплект вентиляции, темно-коричневый/серый")),
    },
    6: {
        "ГК": (10475, note("Wildberries", "GERVENT готовый комплект / комплект кровельной вентиляции, коричневый/черный")),
        "Static/Колпак": (6381, note("Wildberries", "GERVENT комплект кровельной вентиляции Static, коричневый/черный")),
    },
    7: {
        "ГК": (11549, note("Wildberries", "GERVENT комплект с нанодефлектором, серый графит")),
        "РВТ": (9028, note("Wildberries", "GERVENT ротационный нанодефлектор, зеленый")),
        "Static/Колпак": (8790, note("Wildberries", "GERVENT комплект вентиляции Static/колпак, серый")),
    },
    8: {
        "ГК": (13942, note("Wildberries", "GERVENT комплект активной вентиляции, серый графит")),
        "РВТ": (7203, note("Wildberries", "GERVENT ротационный нанодефлектор, серый")),
        "Static/Колпак": (9296, note("Wildberries", "GERVENT комплект вентиляции Static/колпак, серый")),
    },
    13: {
        "ГК": (8645, note("Яндекс Маркет", "Готовый комплект вентиляции Gervent для кровли Монтеррей, черный")),
        "ВВ": (3533, note("Ozon", "Вентиляционный выход изолированный из ABS-пластика ND, серый")),
        "Static/Колпак": (8528, note("Ozon", "Gervent Static комплект для фальцевой/гибкой черепицы, черный")),
    },
    18: {
        "ГК": (11211, note("Ozon", "GERVENT Twister готовый комплект / комплект вентиляционного выхода, серый")),
        "ВВ": (4436, note("Ozon", "Вентиляционный выход изолированный, серый")),
        "Static/Колпак": (6373, note("Ozon", "GERVENT Static готовый комплект, черный/серый")),
    },
    21: {
        "ГК": (11335, note("Wildberries", "GERVENT комплект кровельной вентиляции, серый графит/коричневый")),
        "РВТ": (5842, note("Ozon", "Ротационная вентиляционная турбина Нанодефлектор РВТ-160, коричневый")),
    },
    23: {
        "ГК": (7450, note("Wildberries", "GERVENT комплект для металлочерепицы Монтеррей D160, коричневый")),
    },
    45: {
        "ГК": (11926, note("Ozon", "Вентиляционный выход изолированный для металлочерепицы Монтеррей комплект 150/160, коричневый")),
        "РВТ": (4613, note("Wildberries", "GERVENT ротационная вентиляционная турбина Twister 160 мм, серый графит/коричневый")),
        "ВВ": (4018, note("Ozon", "Вентиляционный выход изолированный Gervent 125/160, серый")),
        "Static/Колпак": (2155, note("Ozon", "Вентиляционный дефлектор Gervent статический 160, серый/коричневый")),
    },
    47: {
        "РВТ": (5976, note("Ozon", "Ротационная вентиляционная турбина 160 Gervent, коричневый RAL 8017")),
        "ВВ": (4122, note("Ozon", "Вентиляционный выход изолированный Gervent 150/160, серый/коричневый")),
    },
    49: {
        "ГК": (11011, note("Ozon", "Комплект вентиляционного выхода на крышу GERVENT Twister 125 мм, черный")),
        "Static/Колпак": (7778, note("Ozon", "GERVENT Static 110 комплект для профнастила С-21, коричневый")),
    },
}


def clone_cell_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)


wb = load_workbook(BASE)
ws = wb["Данные"]

link_col = None
for col in range(1, ws.max_column + 1):
    value = ws.cell(3, col).value
    if isinstance(value, str) and "ссылка" in value.lower():
        link_col = col
        break
if link_col is None:
    link_col = 259

link_snapshot = {}
for col in range(link_col, ws.max_column + 1):
    link_snapshot[col] = {}
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, col)
        link_snapshot[col][row] = {
            "value": cell.value,
            "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "alignment": copy(cell.alignment),
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "comment": copy(cell.comment) if cell.comment else None,
        }

ws.insert_cols(link_col, 4)

new_cols = [link_col + i for i in range(4)]
source_cols = [link_col - 4 + i for i in range(4)]
for src_col, dst_col in zip(source_cols, new_cols):
    ws.column_dimensions[ws.cell(1, dst_col).column_letter].width = ws.column_dimensions[ws.cell(1, src_col).column_letter].width
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, src_col)
        dst = ws.cell(row, dst_col)
        clone_cell_style(src, dst)
        dst.value = None
        dst.hyperlink = None
        dst.comment = None
        if row >= 4:
            dst.fill = copy(NO_FILL)

ws.cell(1, new_cols[0]).value = DATE_TEXT
ws.cell(2, new_cols[0]).value = "Цена, руб"
headers = ["ГК", "РВТ", "ВВ", "Static/Колпак"]
for col, header in zip(new_cols, headers):
    ws.cell(3, col).value = header

our_gk = DATA[4]["ГК"][0]
col_by_header = dict(zip(headers, new_cols))
for row, row_data in DATA.items():
    for header, (price, comment_text) in row_data.items():
        cell = ws.cell(row, col_by_header[header])
        cell.value = price
        cell.number_format = "#,##0"
        cell.comment = Comment(comment_text, "Codex")
        cell.hyperlink = None
        if header == "РВТ" and price < 5750:
            cell.fill = copy(LIGHT_RED)
        elif header == "ГК" and row != 4 and price < our_gk:
            cell.fill = copy(LIGHT_RED)
        elif header == "Static/Колпак" and price < 7900:
            cell.fill = copy(LIGHT_RED)

for old_col, rows in link_snapshot.items():
    new_col = old_col + 4
    for row, data in rows.items():
        cell = ws.cell(row, new_col)
        cell._style = copy(data["style"])
        cell.number_format = data["number_format"]
        cell.alignment = copy(data["alignment"])
        cell.font = copy(data["font"])
        cell.fill = copy(data["fill"])
        cell.border = copy(data["border"])
        cell.value = data["value"]
        cell.hyperlink = copy(data["hyperlink"]) if data["hyperlink"] else None
        cell.comment = copy(data["comment"]) if data["comment"] else None

for col in new_cols:
    for row in range(1, ws.max_row + 1):
        ws.cell(row, col).hyperlink = None

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(OUT)
