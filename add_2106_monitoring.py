from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Users\HONOR\Documents\Codex\2026-06-18\files-mentioned-by-the-user-17062026")
INPUT = BASE / "outputs" / "Обзор цен 19062026 только цены мониторинг авто 1420.xlsx"
OUTPUT = BASE / "outputs" / "Обзор цен 21062026 только цены мониторинг.xlsx"

DATE_TEXT = "21.06.2026 10:45-10:54"
RED_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
NO_FILL = PatternFill(fill_type=None)


def cmt(marketplace: str, title: str) -> str:
    return f"{marketplace}, {DATE_TEXT} — {title}"


data = {
    4: {
        "ГК": (10419, cmt("Wildberries", "GERVENT комплект вентиляции PROF-20/готовый комплект, серый графит")),
        "РВТ": (5248, cmt("Wildberries", "GERVENT Нанодефлектор РВТ-160, серый графит")),
        "Static/Колпак": (7060, cmt("Wildberries", "GERVENT комплект вентиляции Static/колпак, темный цвет")),
    },
    6: {
        "ГК": (11001, cmt("Wildberries", "GERVENT комплект вентиляции / готовый комплект, коричневый/черный")),
        "Static/Колпак": (6707, cmt("Wildberries", "GERVENT комплект вентиляции Static/колпак, коричневый")),
    },
    7: {
        "ГК": (12128, cmt("Wildberries", "GERVENT комплект с нанодефлектором, серый графит")),
        "РВТ": (9481, cmt("Wildberries", "GERVENT ротационный нанодефлектор, зеленый")),
        "Static/Колпак": (8807, cmt("Wildberries", "GERVENT комплект с колпаком универсальный, серый графит")),
    },
    8: {
        "ГК": (14643, cmt("Wildberries", "GERVENT готовый комплект активной вентиляции, серый графит")),
        "РВТ": (7570, cmt("Wildberries", "GERVENT дефлектор ротационный, серый")),
        "Static/Колпак": (9606, cmt("Wildberries", "GERVENT комплект с колпаком, серый графит")),
    },
    13: {
        "ГК": (11706, cmt("Ozon", "GERVENT комплект кровельной вентиляции для металлочерепицы Монтеррей, зеленый")),
        "ВВ": (3427, cmt("Ozon", "GERVENT вентиляционный выход из ABS-пластика ND, серый")),
        "Static/Колпак": (5947, cmt("Wildberries", "GERVENT комплект вентиляции Static, черный/коричневый")),
    },
    18: {
        "ГК": (9977, cmt("Ozon", "Готовый комплект изолированный для кровли, серия Twister, черный RAL 9005")),
        "РВТ": (3138, cmt("Яндекс Маркет", "Ротационная вентиляционная турбина РВТ 160, серебристый/бордовый")),
        "ВВ": (2088, cmt("Яндекс Маркет", "Вентиляционный выход изолированный 110/160 мм, коричневый/серый графит")),
        "Static/Колпак": (6884, cmt("Ozon", "Готовый комплект изолированный для кровли, серия Static, черный RAL 9005")),
    },
    21: {
        "ГК": (11225, cmt("Ozon", "Комплект вентиляции для плоской кровли, вытяжка для бани, серый графит")),
        "РВТ": (5842, cmt("Ozon", "Ротационная вентиляционная турбина Нанодефлектор РВТ-160, коричневый")),
    },
    23: {
        "Static/Колпак": (7450, cmt("Wildberries", "GERVENT комплект для металлочерепицы Монтеррей, коричневый")),
    },
    45: {
        "РВТ": (4540, cmt("Wildberries", "GERVENT ротационная вентиляционная турбина, цвет не прогружен на карточке")),
        "ВВ": (4141, cmt("Ozon", "GERVENT вентиляционный выход изолированный 150/160, серый RAL 7024")),
    },
    47: {
        "РВТ": (5959, cmt("Ozon", "GERVENT ротационная вентиляционная турбина 160, серый/коричневый")),
        "ВВ": (4105, cmt("Ozon", "GERVENT вентиляционный выход изолированный 150/160, серый/коричневый")),
    },
    49: {
        "ГК": (10463, cmt("Ozon", "GERVENT Twister 150 комплект для профнастила C-21, черный")),
        "Static/Колпак": (7504, cmt("Ozon", "GERVENT Static 110 комплект для профнастила C-21, коричневый")),
    },
}


wb = load_workbook(INPUT)
ws = wb.active

link_col = None
for col in range(1, ws.max_column + 1):
    value = ws.cell(3, col).value
    if isinstance(value, str) and "ссылка" in value.lower():
        link_col = col
        break

if link_col is None:
    raise RuntimeError("Не удалось найти колонки со ссылками.")

source_cols = list(range(link_col - 4, link_col))
ws.insert_cols(link_col, 4)
new_cols = list(range(link_col, link_col + 4))

for src_col, dst_col in zip(source_cols, new_cols):
    ws.column_dimensions[get_column_letter(dst_col)].width = ws.column_dimensions[get_column_letter(src_col)].width
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, src_col)
        dst = ws.cell(row, dst_col)
        dst.value = None
        dst.hyperlink = None
        dst.comment = None
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.value = None
        dst.hyperlink = None
        dst.comment = None

for col in new_cols:
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, col)
        cell.value = None
        cell.hyperlink = None
        cell.comment = None

headers = ["ГК", "РВТ", "ВВ", "Static/Колпак"]
ws.cell(1, new_cols[0]).value = "21.06.2026"
ws.cell(2, new_cols[0]).value = "Цена, руб"
for col, header in zip(new_cols, headers):
    ws.cell(3, col).value = header
for col in new_cols[1:]:
    ws.cell(1, col).value = None
    ws.cell(2, col).value = None

for row, row_data in data.items():
    for col, header in zip(new_cols, headers):
        cell = ws.cell(row, col)
        cell.value = None
        cell.hyperlink = None
        cell.comment = None
        cell.fill = copy(NO_FILL)
        if header not in row_data:
            continue
        value, comment = row_data[header]
        cell.value = value
        cell.comment = Comment(comment, "Codex")
        if header == "РВТ" and value < 5750:
            cell.fill = copy(RED_FILL)
        elif header == "Static/Колпак" and value < 7900:
            cell.fill = copy(RED_FILL)

own_gk = data[4]["ГК"][0]
for row, row_data in data.items():
    if row == 4 or "ГК" not in row_data:
        continue
    value, _ = row_data["ГК"]
    if value < own_gk:
        ws.cell(row, new_cols[0]).fill = copy(RED_FILL)

wb.save(OUTPUT)
print(OUTPUT)
