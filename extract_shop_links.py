import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


DOMAINS = {
    "ozon.ru": "Ozon",
    "wildberries.ru": "Wildberries",
    "market.yandex.ru": "YandexMarket",
}


def marketplace_from_url(url: str, fallback: str = "") -> str:
    text = (url or "").lower()
    for domain, name in DOMAINS.items():
        if domain in text:
            return name
    fallback = (fallback or "").lower()
    if "ozon" in fallback:
        return "Ozon"
    if "wb" in fallback or "wildberries" in fallback:
        return "Wildberries"
    if "яндекс" in fallback or "yandex" in fallback:
        return "YandexMarket"
    return "Marketplace"


def first_url(value: str) -> str | None:
    if not value:
        return None
    match = re.search(r"https?://\S+", str(value))
    if not match:
        return None
    return match.group(0).rstrip(".,);")


def main() -> None:
    if len(sys.argv) < 3:
        raise SystemExit("Usage: extract_shop_links.py <workbook.xlsx> <links.json>")

    workbook_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    wb = load_workbook(workbook_path, data_only=False)
    ws = wb.active

    link_cols = []
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(3, col).value or "").lower()
        if "ссылка" in header and "магазин" in header:
            link_cols.append(col)

    if not link_cols:
        for col in range(1, ws.max_column + 1):
            found = False
            for row in range(4, ws.max_row + 1):
                cell = ws.cell(row, col)
                url = cell.hyperlink.target if cell.hyperlink else first_url(cell.value)
                if url and any(domain in url.lower() for domain in DOMAINS):
                    found = True
                    break
            if found:
                link_cols.append(col)

    links = []
    seen = set()
    for row in range(4, ws.max_row + 1):
        seller = str(ws.cell(row, 1).value or "").strip()
        if not seller:
            continue
        for col in link_cols:
            cell = ws.cell(row, col)
            url = cell.hyperlink.target if cell.hyperlink else first_url(cell.value)
            if not url:
                continue
            if not any(domain in url.lower() for domain in DOMAINS):
                continue
            header = str(ws.cell(3, col).value or "")
            marketplace = marketplace_from_url(url, header)
            key = (row, marketplace, url)
            if key in seen:
                continue
            seen.add(key)
            links.append(
                {
                    "row": row,
                    "seller": seller,
                    "marketplace": marketplace,
                    "url": url,
                    "display": str(cell.value or url),
                }
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(links, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Saved {len(links)} links to {output_path}")


if __name__ == "__main__":
    main()
